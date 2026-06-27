from pathlib import Path
import shutil

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
source = (
    ROOT
    / "outputs"
    / "GeradorMemorialSolar"
    / "assets"
    / "Mapa_Dados_Memorial_LIGHT_100_PARAM.xlsx"
)
target = ROOT / "work" / "Mapa_Dados_Teste_Sonia.xlsx"
shutil.copy2(source, target)

values = {
    "CLIENTE_NOME": "SONIA MEIRELLES ALMEIDA",
    "CLIENTE_ENDERECO_LOGRADOURO": "R JOSE DOS REIS 47 LJ",
    "CLIENTE_BAIRRO_MUN_UF_CEP": "COELHO DA ROCHA / SAO JOAO DE MERITI, RJ, CEP 25550-780",
    "LOCAL_CIDADE_UF": "SAO JOAO DE MERITI, RJ",
    "MES": "JUNHO",
    "ANO": "2026",
    "UC_CONTA_CONTRATO": "2.783.802.059-28",
    "UC_CLASSE": "B3",
    "COORD_SUL": -22.7818341,
    "COORD_OESTE": -43.3807375,
    "FV_POT_KWP": 5.85,
    "MOD_MARCA": "RONMA",
    "MOD_MODELO": "RM-585W-182M/144TB",
    "MOD_QTD": 10,
    "MOD_WP": 585,
    "MOD_EFIC": 22.6,
    "MOD_VMP": 43.27,
    "MOD_IMP": 13.52,
    "MOD_VOC": 51.50,
    "MOD_ISC": 14.36,
    "MOD_VSYS_MAX": 1500,
    "MOD_FUSIVEL_MAX": 30,
    "MOD_TIPO_CELULA": "N-TOPCon Mono bifacial",
    "MOD_N_CELULAS": 144,
    "MOD_DIMENSOES": "2278 x 1134 x 35 mm",
    "MOD_PESO_KG": 32,
    "INV_MARCA": "DEYE",
    "INV_MODELO": "SUN-4K-G04",
    "INV_QTD": 1,
    "INV_PN_KW": 4,
    "INV_PMAXCC_KW": 5.2,
    "INV_VCC_MAX": 550,
    "INV_ICC_MAX": 39,
    "INV_VMPPT_MAX": 500,
    "INV_VMPPT_MIN": 70,
    "INV_VSTART": 80,
    "INV_STRINGS": 2,
    "INV_MPPTS": 2,
    "INV_PCA_KW": 4,
    "INV_PCA_MAX_KW": 4.4,
    "INV_IMAX_CA": 20,
    "INV_VAC_NOM": 220,
    "INV_FN": 60,
    "INV_FP": "0,8 adiantado a 0,8 atrasado",
    "INV_CONEXAO": "L/N/PE",
    "INV_VCA_MAX": 242,
    "INV_VCA_MIN": 187,
    "INV_THD": 3,
    "INV_EFIC_MAX": 97.5,
    "DIST_CC_M": 10,
    "DIST_CA_M": 10,
    "BITOLA_CC_MM2": 6,
    "BITOLA_CA_MM2": 6,
    "PROT_CC": 20,
    "PROT_CA": 32,
    "R_CC_OHMKM": 4.32,
    "R_CA_OHMKM": 4.32,
}

workbook = load_workbook(target)
inputs = workbook["Inputs"]
row_by_key = {
    inputs.cell(row, 1).value: row
    for row in range(2, inputs.max_row + 1)
}
for key, value in values.items():
    inputs.cell(row_by_key[key], 2).value = value

consumption = [302, 281, 40, 37, 34, 31, 0, 0, 0, 0, 0, 0]
for index, value in enumerate(consumption, start=2):
    workbook["Consumo_24m"].cell(index, 2).value = value

workbook.save(target)
