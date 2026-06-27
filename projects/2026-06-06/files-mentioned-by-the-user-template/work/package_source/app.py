from __future__ import annotations

import shutil
import sys
import tkinter as tk
from pathlib import Path
from tkinter import filedialog, messagebox, ttk

from openpyxl import load_workbook

from memorial import generate_memorial
from memorial.extractors import suggest_fields


def resource_path(name: str) -> Path:
    base = Path(getattr(sys, "_MEIPASS", Path(__file__).parent))
    return base / "assets" / name


def ensure_workbook_schema(user_excel: Path, template_excel: Path) -> None:
    user_wb = load_workbook(user_excel)
    template_wb = load_workbook(template_excel)

    user_inputs = user_wb["Inputs"]
    template_inputs = template_wb["Inputs"]
    existing_inputs = {user_inputs[f"A{row}"].value for row in range(2, user_inputs.max_row + 1)}
    for row in range(2, template_inputs.max_row + 1):
        key = template_inputs[f"A{row}"].value
        if key and key not in existing_inputs:
            dest = user_inputs.max_row + 1
            user_inputs[f"A{dest}"] = key
            user_inputs[f"B{dest}"] = template_inputs[f"B{row}"].value
            existing_inputs.add(key)

    user_map = user_wb["Mapeamento"]
    template_map = template_wb["Mapeamento"]
    existing_map = {user_map[f"A{row}"].value for row in range(2, user_map.max_row + 1)}
    for row in range(2, template_map.max_row + 1):
        placeholder = template_map[f"A{row}"].value
        if placeholder and placeholder not in existing_map:
            dest = user_map.max_row + 1
            user_map[f"A{dest}"] = placeholder
            user_map[f"B{dest}"] = template_map[f"B{row}"].value
            existing_map.add(placeholder)
    user_wb.save(user_excel)


FIELD_GROUPS = [
    ("Cliente e UC", [
        "CLIENTE_NOME", "CLIENTE_ENDERECO_LOGRADOURO", "CLIENTE_BAIRRO_MUN_UF_CEP",
        "LOCAL_CIDADE_UF", "MES", "ANO", "UC_CONTA_CONTRATO", "UC_CLASSE",
        "COORD_SUL", "COORD_OESTE", "CONCESSIONARIA", "ESTADO",
    ]),
    ("Sistema FV", [
        "FV_POT_KWP", "PERDAS_GERACAO", "TIPO_LIGACAO", "CONDUTOR_FASE_MM2",
        "CONDUTOR_NEUTRO_DESC", "DISJ_POLOS", "DISJ_CORRENTE_A", "DISJ_TENSAO_V",
    ]),
    ("Datasheet Modulo", [
        "MOD_MARCA", "MOD_MODELO", "MOD_QTD", "MOD_WP", "MOD_EFIC", "MOD_VMP",
        "MOD_IMP", "MOD_VOC", "MOD_ISC", "MOD_VSYS_MAX", "MOD_FUSIVEL_MAX",
        "MOD_TIPO_CELULA", "MOD_N_CELULAS", "MOD_DIMENSOES", "MOD_PESO_KG",
    ]),
    ("Datasheet Inversor", [
        "INV_MARCA", "INV_MODELO", "INV_QTD", "INV_PN_KW", "INV_PMAXCC_KW",
        "INV_VCC_MAX", "INV_ICC_MAX", "INV_VMPPT_MAX", "INV_VMPPT_MIN",
        "INV_VSTART", "INV_STRINGS", "INV_MPPTS", "INV_PCA_KW", "INV_PCA_MAX_KW",
        "INV_IMAX_CA", "INV_VAC_NOM", "INV_FN", "INV_FP", "INV_CONEXAO",
        "INV_VCA_MAX", "INV_VCA_MIN", "INV_THD", "INV_EFIC_MAX",
    ]),
    ("Protecao, cabos e aterramento", [
        "DIST_CC_M", "DIST_CA_M", "BITOLA_CC_MM2", "BITOLA_CA_MM2", "PROT_CC",
        "PROT_CA", "R_CC_OHMKM", "R_CA_OHMKM", "PROT_DISJ_POLOS",
        "PROT_DISJ_TENSAO_V", "PROT_DISJ_CORRENTE_A", "PROT_DISJ_FREQ_HZ",
        "PROT_DISJ_CAP_INT_KA", "PROT_DISJ_CURVA", "DSV_POLOS", "DSV_TENSAO_V",
        "DSV_CORRENTE_A", "DSV_FECHAMENTO_CC_A", "DSV_CAP_INT_KA", "DSV_CURVA",
        "ATERRAMENTO_ESQUEMA", "ATERRAMENTO_RESISTENCIA_OHM", "ATERRAMENTO_HASTES_QTD",
        "ATERRAMENTO_HASTE_M", "ATERRAMENTO_HASTE_SECAO", "ATERRAMENTO_BITOLA_MM2",
        "CABO_ISOLACAO", "CABO_ISOLAMENTO_KV", "CABO_BITOLA_MM2", "CABO_CAPACIDADE_A",
        "CABO_METODO_INSTALACAO", "FATOR_TEMPERATURA", "FATOR_AGRUPAMENTO",
        "TEMPERATURA_AMBIENTE_C",
    ]),
]


class MemorialApp(tk.Tk):
    def __init__(self) -> None:
        super().__init__()
        self.title("Gerador de Memorial Solar")
        self.geometry("1180x780")
        self.minsize(980, 650)

        user_dir = Path.home() / "Documents" / "Gerador Memorial Solar"
        user_dir.mkdir(parents=True, exist_ok=True)
        user_excel = user_dir / "Mapa_Dados_Memorial_LIGHT_100_PARAM.xlsx"
        template_excel = resource_path("Mapa_Dados_Memorial_LIGHT_100_PARAM.xlsx")
        if not user_excel.exists():
            shutil.copy2(template_excel, user_excel)
        else:
            ensure_workbook_schema(user_excel, template_excel)

        self.excel = tk.StringVar(value=str(user_excel))
        self.template = tk.StringVar(value=str(resource_path("TEMPLATE_MEMORIAL_LIGHT_100_PARAM_TAGUEADO.docx")))
        self.output = tk.StringVar(value=str(user_dir / "Memorial_Gerado.docx"))
        self.documents: list[str] = []
        self.consumption_suggestions: list[int] = []
        self.fields: dict[str, tk.StringVar] = {}
        self.consumption_fields: dict[int, tk.StringVar] = {}
        self.load_fields: dict[tuple[int, str], tk.StringVar] = {}
        self._build()
        self.load_excel()

    def _path_row(self, parent, row: int, label: str, variable: tk.StringVar, command) -> None:
        ttk.Label(parent, text=label).grid(row=row, column=0, sticky="w", padx=6, pady=4)
        ttk.Entry(parent, textvariable=variable).grid(row=row, column=1, sticky="ew", padx=6, pady=4)
        ttk.Button(parent, text="Selecionar", command=command).grid(row=row, column=2, padx=6, pady=4)

    def _scroll_tab(self, title: str) -> ttk.Frame:
        outer = ttk.Frame(self.notebook)
        self.notebook.add(outer, text=title)
        canvas = tk.Canvas(outer, highlightthickness=0)
        scrollbar = ttk.Scrollbar(outer, orient="vertical", command=canvas.yview)
        frame = ttk.Frame(canvas)
        frame.bind("<Configure>", lambda event: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.create_window((0, 0), window=frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        return frame

    def _build(self) -> None:
        top = ttk.LabelFrame(self, text="Arquivos do projeto")
        top.pack(fill="x", padx=12, pady=10)
        top.columnconfigure(1, weight=1)
        self._path_row(top, 0, "Planilha", self.excel, self.choose_excel)
        self._path_row(top, 1, "Template Word", self.template, self.choose_template)
        self._path_row(top, 2, "Memorial de saida", self.output, self.choose_output)

        actions = ttk.Frame(self)
        actions.pack(fill="x", padx=12)
        ttk.Button(actions, text="Adicionar documentos", command=self.choose_documents).pack(side="left", padx=4)
        ttk.Button(actions, text="Analisar documentos", command=self.analyze_documents).pack(side="left", padx=4)
        ttk.Button(actions, text="Recarregar planilha", command=self.load_excel).pack(side="left", padx=4)
        ttk.Button(actions, text="Salvar dados", command=self.save_excel).pack(side="left", padx=4)
        ttk.Button(actions, text="Gerar memorial", command=self.generate).pack(side="right", padx=4)

        self.status = tk.StringVar(
            value="Use PDFs legiveis/texto selecionavel. Datasheets e conta escaneados precisam de revisao manual."
        )
        ttk.Label(self, textvariable=self.status, foreground="#204a87").pack(fill="x", padx=16, pady=8)

        self.notebook = ttk.Notebook(self)
        self.notebook.pack(fill="both", expand=True, padx=12, pady=8)
        self.inputs_frame = self._scroll_tab("Formulario")
        self.consumption_frame = self._scroll_tab("Consumo 12 meses")
        self.loads_frame = self._scroll_tab("Cargas da residencia")

    def choose_excel(self) -> None:
        path = filedialog.askopenfilename(filetypes=[("Excel", "*.xlsx")])
        if path:
            self.excel.set(path)
            self.load_excel()

    def choose_template(self) -> None:
        path = filedialog.askopenfilename(filetypes=[("Word", "*.docx")])
        if path:
            self.template.set(path)

    def choose_output(self) -> None:
        path = filedialog.asksaveasfilename(defaultextension=".docx", filetypes=[("Word", "*.docx")])
        if path:
            self.output.set(path)

    def choose_documents(self) -> None:
        paths = filedialog.askopenfilenames(
            filetypes=[("Documentos", "*.pdf *.docx *.txt *.csv *.png *.jpg *.jpeg"), ("Todos", "*.*")]
        )
        self.documents = list(paths)
        self.status.set(f"{len(self.documents)} documento(s) selecionado(s).")

    def _clear_frames(self) -> None:
        for frame in (self.inputs_frame, self.consumption_frame, self.loads_frame):
            for widget in frame.winfo_children():
                widget.destroy()

    def load_excel(self) -> None:
        try:
            workbook = load_workbook(self.excel.get(), data_only=False)
        except Exception as exc:
            messagebox.showerror("Planilha", str(exc))
            return

        self._clear_frames()
        self.fields.clear()
        self.consumption_fields.clear()
        self.load_fields.clear()
        inputs = workbook["Inputs"]
        input_values = {
            str(inputs[f"A{row}"].value or ""): inputs[f"B{row}"].value
            for row in range(2, inputs.max_row + 1)
        }

        display_row = 0
        used: set[str] = set()
        for group, keys in FIELD_GROUPS:
            ttk.Label(self.inputs_frame, text=group, font=("Segoe UI", 10, "bold")).grid(
                row=display_row, column=0, columnspan=2, sticky="w", padx=6, pady=(10, 4)
            )
            display_row += 1
            for key in keys:
                value = input_values.get(key)
                variable = tk.StringVar(value="" if value is None else str(value))
                self.fields[key] = variable
                used.add(key)
                ttk.Label(self.inputs_frame, text=key, width=36).grid(row=display_row, column=0, sticky="w", padx=6, pady=3)
                ttk.Entry(self.inputs_frame, textvariable=variable, width=72).grid(
                    row=display_row, column=1, sticky="ew", padx=6, pady=3
                )
                display_row += 1
        for key, value in input_values.items():
            if key and key not in used:
                variable = tk.StringVar(value="" if value is None else str(value))
                self.fields[key] = variable
                ttk.Label(self.inputs_frame, text=key, width=36).grid(row=display_row, column=0, sticky="w", padx=6, pady=3)
                ttk.Entry(self.inputs_frame, textvariable=variable, width=72).grid(
                    row=display_row, column=1, sticky="ew", padx=6, pady=3
                )
                display_row += 1
        self.inputs_frame.columnconfigure(1, weight=1)

        consumo = workbook["Consumo_24m"]
        ttk.Label(self.consumption_frame, text="Preencha os 12 meses da conta de energia", font=("Segoe UI", 10, "bold")).grid(
            row=0, column=0, columnspan=2, sticky="w", padx=6, pady=(10, 4)
        )
        for idx in range(1, 13):
            value = consumo[f"B{idx + 1}"].value
            variable = tk.StringVar(value="" if value is None else str(value))
            self.consumption_fields[idx] = variable
            ttk.Label(self.consumption_frame, text=f"Consumo mes {idx} (kWh)", width=28).grid(
                row=idx, column=0, sticky="w", padx=6, pady=3
            )
            ttk.Entry(self.consumption_frame, textvariable=variable, width=24).grid(
                row=idx, column=1, sticky="w", padx=6, pady=3
            )

        cargas = workbook["Cargas_24"]
        headers = [("DESC", "Descricao"), ("P_W", "Potencia W"), ("QTD", "Quantidade"), ("FP", "FP")]
        for col, (_, label) in enumerate(headers, start=1):
            ttk.Label(self.loads_frame, text=label, font=("Segoe UI", 9, "bold")).grid(row=0, column=col, padx=6, pady=4)
        for idx in range(1, 13):
            ttk.Label(self.loads_frame, text=f"Carga {idx}").grid(row=idx, column=0, sticky="w", padx=6, pady=3)
            for col_offset, (suffix, _) in enumerate(headers, start=2):
                value = cargas.cell(idx + 1, col_offset).value
                variable = tk.StringVar(value="" if value is None else str(value))
                self.load_fields[(idx, suffix)] = variable
                ttk.Entry(self.loads_frame, textvariable=variable, width=24).grid(
                    row=idx, column=col_offset - 1, sticky="ew", padx=4, pady=3
                )

    def analyze_documents(self) -> None:
        if not self.documents:
            messagebox.showwarning("Documentos", "Adicione os PDFs ou documentos antes de analisar.")
            return
        try:
            suggestions, unsupported, consumption = suggest_fields(self.documents)
        except Exception as exc:
            messagebox.showerror("Analise", str(exc))
            return
        for key, value in suggestions.items():
            if key in self.fields and not self.fields[key].get().strip():
                self.fields[key].set(value)
        self.consumption_suggestions = consumption
        for idx, value in enumerate(consumption[:12], start=1):
            if idx in self.consumption_fields and not self.consumption_fields[idx].get().strip():
                self.consumption_fields[idx].set(str(value))
        note = f"{len(suggestions)} campo(s) sugerido(s)."
        if consumption:
            note += f" {len(consumption)} consumo(s) mensais detectado(s)."
        if unsupported:
            note += " Sem OCR para: " + ", ".join(unsupported)
        self.status.set(note + " Revise tudo antes de gerar.")

    def save_excel(self) -> None:
        try:
            workbook = load_workbook(self.excel.get(), data_only=False)
            inputs = workbook["Inputs"]
            for row in range(2, inputs.max_row + 1):
                key = str(inputs[f"A{row}"].value or "")
                if key in self.fields:
                    inputs[f"B{row}"] = self.fields[key].get().strip() or None

            consumo = workbook["Consumo_24m"]
            for idx, variable in self.consumption_fields.items():
                consumo[f"B{idx + 1}"] = variable.get().strip() or None

            cargas = workbook["Cargas_24"]
            column_by_suffix = {"DESC": 2, "P_W": 3, "QTD": 4, "FP": 5}
            for (idx, suffix), variable in self.load_fields.items():
                cargas.cell(idx + 1, column_by_suffix[suffix]).value = variable.get().strip() or None

            workbook.save(self.excel.get())
            self.status.set("Dados salvos na planilha.")
        except Exception as exc:
            messagebox.showerror("Salvar", str(exc))

    def generate(self) -> None:
        self.save_excel()
        try:
            result = generate_memorial(self.excel.get(), self.template.get(), self.output.get())
        except Exception as exc:
            messagebox.showerror("Geracao", str(exc))
            return
        details = [f"Memorial gerado em:\n{result.output}"]
        if result.missing_values:
            details.append(f"\nAtencao: {len(result.missing_values)} campos estavam vazios.")
        if result.unresolved_tags:
            details.append("\nTags nao resolvidas:\n" + ", ".join(result.unresolved_tags))
        messagebox.showinfo("Concluido", "\n".join(details))
        self.status.set("Geracao concluida. Revise tecnicamente o documento antes de protocolar.")


if __name__ == "__main__":
    MemorialApp().mainloop()
