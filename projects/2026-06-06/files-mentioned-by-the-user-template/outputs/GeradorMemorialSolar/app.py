from __future__ import annotations

import json
import shutil
import sqlite3
import sys
import tkinter as tk
from datetime import date, datetime
from pathlib import Path
from tkinter import filedialog, messagebox, ttk

from openpyxl import load_workbook

from memorial import (
    AllocationAccount,
    generate_concessionaire_package,
    generate_memorial,
    generate_single_line_diagram,
)
from memorial.extractors import suggest_fields
from memorial.shared_data import expand_shared_values


EXTRA_INPUT_DEFAULTS = {
    "PROJETO_NOME": None,
    "PROJETO_STATUS": "Em preenchimento",
    "COTACAO_NUMERO": None,
    "STRING_QTD_SUGERIDA": None,
    "MOD_POR_STRING_SUGERIDO": None,
    "CLIENTE_CPF": None,
    "CLIENTE_RG": None,
    "CLIENTE_DATA_NASCIMENTO": None,
    "CLIENTE_NATURALIDADE": None,
    "CLIENTE_EMAIL": None,
    "CLIENTE_CELULAR": None,
    "CLIENTE_TELEFONE": None,
    "UC_CODIGO_CLIENTE": None,
    "CARGA_INSTALADA_KW": None,
    "TIPO_RAMAL": "Aéreo",
    "TIPO_SOLICITACAO": "Conexão em UC existente sem alteração de potência",
    "DATA_IMPLEMENTACAO": None,
    "DATA_CONEXAO": None,
    "CEG_EMPREENDIMENTO": None,
    "USINA_NOME": None,
    "OUTORGA_TIPO": None,
    "OUTORGA_NUMERO": None,
    "OUTORGA_ANO": None,
    "SOLICITANTE_NOME": None,
    "RT_NOME": None,
    "RT_REGISTRO": None,
    "RT_ART_TRT": None,
    "RT_EMAIL": None,
    "RT_CELULAR": None,
    "RT_CPF": None,
    "DATA_PROJETO": None,
    "INMETRO_REGISTRO": None,
    "INMETRO_DATA": None,
    "INV_DIMENSOES": None,
    "INV_PESO_KG": None,
    "INV_IP": None,
    "INV_CLASSE_PROTECAO": None,
    "INV_CONSUMO_NOTURNO_W": None,
    "INV_TOPOLOGIA": None,
    "INV_REFRIGERACAO": None,
    "INV_TEMPERATURA": None,
    "INV_UMIDADE": None,
    "DIAG_DISJ_INV_A": None,
}


def resource_path(name: str) -> Path:
    base = Path(getattr(sys, "_MEIPASS", Path(__file__).parent))
    return base / "assets" / name


MEMORIAL_TEMPLATES = {
    "LIGHT": "TEMPLATE_MEMORIAL_LIGHT_100_PARAM_TAGUEADO.docx",
    "Enel-RJ": "TEMPLATE_MEMORIAL_ENEL_RJ_100_PARAM_TAGUEADO.docx",
}

DIAGRAM_TEMPLATES = {
    "LIGHT": "TEMPLATE_DIAGRAMA_LIGHT.dxf",
    "Enel-RJ": "TEMPLATE_DIAGRAMA_ENEL_RJ.dxf",
}


def ensure_workbook_schema(user_excel: Path, template_excel: Path) -> None:
    user_wb = load_workbook(user_excel)
    template_wb = load_workbook(template_excel)

    user_inputs = user_wb["Inputs"]
    template_inputs = template_wb["Inputs"]
    existing_inputs = {user_inputs[f"A{row}"].value for row in range(2, user_inputs.max_row + 1)}
    for row in range(2, template_inputs.max_row + 1):
        key = template_inputs[f"A{row}"].value
        if key and key not in existing_inputs:
            dest = user_inputs.max_row + 1
            user_inputs[f"A{dest}"] = key
            user_inputs[f"B{dest}"] = template_inputs[f"B{row}"].value
            existing_inputs.add(key)
    for key, value in EXTRA_INPUT_DEFAULTS.items():
        if key not in existing_inputs:
            dest = user_inputs.max_row + 1
            user_inputs[f"A{dest}"] = key
            user_inputs[f"B{dest}"] = value
            existing_inputs.add(key)

    user_map = user_wb["Mapeamento"]
    template_map = template_wb["Mapeamento"]
    existing_map = {user_map[f"A{row}"].value for row in range(2, user_map.max_row + 1)}
    for row in range(2, template_map.max_row + 1):
        placeholder = template_map[f"A{row}"].value
        if placeholder and placeholder not in existing_map:
            dest = user_map.max_row + 1
            user_map[f"A{dest}"] = placeholder
            user_map[f"B{dest}"] = template_map[f"B{row}"].value
            existing_map.add(placeholder)
    user_wb.save(user_excel)




def _safe_filename(value: str, fallback: str = "projeto") -> str:
    cleaned = "".join(char if char.isalnum() else "_" for char in value.strip())
    cleaned = "_".join(part for part in cleaned.split("_") if part)
    return (cleaned or fallback)[:80]


def _project_label(values: dict[str, str]) -> str:
    return (
        values.get("PROJETO_NOME")
        or values.get("CLIENTE_NOME")
        or values.get("UC_CONTA_CONTRATO")
        or "Projeto sem nome"
    ).strip()


def _backup_workbook(excel_path: Path, project_name: str) -> Path | None:
    if not excel_path.exists():
        return None
    backup_dir = excel_path.parent / "Backups"
    backup_dir.mkdir(parents=True, exist_ok=True)
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    target = backup_dir / f"{stamp}_{_safe_filename(project_name)}.xlsx"
    shutil.copy2(excel_path, target)
    return target


def _ensure_history_sheet(workbook) -> None:
    if "Historico_Projetos" not in workbook.sheetnames:
        sheet = workbook.create_sheet("Historico_Projetos")
        sheet.append([
            "Data",
            "Projeto",
            "Cliente",
            "CPF",
            "UC",
            "Concessionaria",
            "kWp",
            "Modulo",
            "Qtd modulos",
            "Inversor",
            "ART",
            "Cotacao",
        ])


def _append_history_row(workbook, values: dict[str, str]) -> None:
    _ensure_history_sheet(workbook)
    sheet = workbook["Historico_Projetos"]
    row = [
        datetime.now().strftime("%d/%m/%Y %H:%M"),
        _project_label(values),
        values.get("CLIENTE_NOME", ""),
        values.get("CLIENTE_CPF", ""),
        values.get("UC_CONTA_CONTRATO", ""),
        values.get("CONCESSIONARIA", ""),
        values.get("FV_POT_KWP", ""),
        values.get("MOD_MODELO", ""),
        values.get("MOD_QTD", ""),
        values.get("INV_MODELO", ""),
        values.get("RT_ART_TRT", ""),
        values.get("COTACAO_NUMERO", ""),
    ]
    existing_key = tuple(str(value or "") for value in row[1:])
    for idx in range(sheet.max_row, 1, -1):
        current_key = tuple(str(sheet.cell(idx, col).value or "") for col in range(2, 13))
        if current_key == existing_key:
            sheet.delete_rows(idx)
            break
    sheet.append(row)


def _save_project_database(db_path: Path, values: dict[str, str], consumption: dict[int, str], loads: dict[tuple[int, str], str], documents: list[str]) -> None:
    db_path.parent.mkdir(parents=True, exist_ok=True)
    with sqlite3.connect(db_path) as conn:
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS projects (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                saved_at TEXT NOT NULL,
                project_name TEXT,
                client_name TEXT,
                client_cpf TEXT,
                uc TEXT,
                concessionaire TEXT,
                payload_json TEXT NOT NULL
            )
            """
        )
        payload = {
            "fields": values,
            "consumption": consumption,
            "loads": {f"{row}_{suffix}": value for (row, suffix), value in loads.items()},
            "documents": documents,
        }
        conn.execute(
            """
            INSERT INTO projects (saved_at, project_name, client_name, client_cpf, uc, concessionaire, payload_json)
            VALUES (?, ?, ?, ?, ?, ?, ?)
            """,
            (
                datetime.now().isoformat(timespec="seconds"),
                _project_label(values),
                values.get("CLIENTE_NOME", ""),
                values.get("CLIENTE_CPF", ""),
                values.get("UC_CONTA_CONTRATO", ""),
                values.get("CONCESSIONARIA", ""),
                json.dumps(payload, ensure_ascii=False),
            ),
        )


FIELD_GROUPS = [
    ("Projeto", [
        "PROJETO_NOME", "PROJETO_STATUS", "COTACAO_NUMERO",
        "STRING_QTD_SUGERIDA", "MOD_POR_STRING_SUGERIDO",
    ]),
    ("Cliente e UC", [
        "CLIENTE_NOME", "CLIENTE_CPF", "CLIENTE_RG", "CLIENTE_DATA_NASCIMENTO",
        "CLIENTE_NATURALIDADE", "CLIENTE_ENDERECO_LOGRADOURO",
        "CLIENTE_BAIRRO_MUN_UF_CEP", "LOCAL_CIDADE_UF", "MES", "ANO",
        "UC_CONTA_CONTRATO", "UC_CLASSE", "COORD_SUL", "COORD_OESTE",
        "CONCESSIONARIA", "ESTADO", "CLIENTE_EMAIL", "CLIENTE_CELULAR",
        "CLIENTE_TELEFONE", "UC_CODIGO_CLIENTE",
    ]),
    ("Sistema FV", [
        "FV_POT_KWP", "PERDAS_GERACAO", "TIPO_LIGACAO", "CONDUTOR_FASE_MM2",
        "CONDUTOR_NEUTRO_DESC", "DISJ_POLOS", "DISJ_CORRENTE_A", "DISJ_TENSAO_V",
        "CARGA_INSTALADA_KW", "TIPO_RAMAL", "TIPO_SOLICITACAO", "DATA_IMPLEMENTACAO",
        "DATA_CONEXAO", "CEG_EMPREENDIMENTO", "USINA_NOME", "OUTORGA_TIPO",
        "OUTORGA_NUMERO", "OUTORGA_ANO",
    ]),
    ("Solicitante e responsável técnico", [
        "SOLICITANTE_NOME", "RT_NOME", "RT_CPF", "RT_REGISTRO", "RT_ART_TRT",
        "RT_EMAIL", "RT_CELULAR", "DATA_PROJETO",
    ]),
    ("Datasheet Modulo", [
        "MOD_MARCA", "MOD_MODELO", "MOD_QTD", "MOD_WP", "MOD_EFIC", "MOD_VMP",
        "MOD_IMP", "MOD_VOC", "MOD_ISC", "MOD_VSYS_MAX", "MOD_FUSIVEL_MAX",
        "MOD_TIPO_CELULA", "MOD_N_CELULAS", "MOD_DIMENSOES", "MOD_PESO_KG",
    ]),
    ("Datasheet Inversor", [
        "INV_MARCA", "INV_MODELO", "INV_QTD", "INV_PN_KW", "INV_PMAXCC_KW",
        "INV_VCC_MAX", "INV_ICC_MAX", "INV_VMPPT_MAX", "INV_VMPPT_MIN",
        "INV_VSTART", "INV_STRINGS", "INV_MPPTS", "INV_PCA_KW", "INV_PCA_MAX_KW",
        "INV_IMAX_CA", "INV_VAC_NOM", "INV_FN", "INV_FP", "INV_CONEXAO",
        "INV_VCA_MAX", "INV_VCA_MIN", "INV_THD", "INV_EFIC_MAX",
        "INMETRO_REGISTRO", "INMETRO_DATA", "INV_DIMENSOES", "INV_PESO_KG",
        "INV_IP", "INV_CLASSE_PROTECAO", "INV_CONSUMO_NOTURNO_W",
        "INV_TOPOLOGIA", "INV_REFRIGERACAO", "INV_TEMPERATURA", "INV_UMIDADE",
    ]),
    ("Protecao, cabos e aterramento", [
        "DIST_CC_M", "DIST_CA_M", "BITOLA_CC_MM2", "BITOLA_CA_MM2", "PROT_CC",
        "PROT_CA", "R_CC_OHMKM", "R_CA_OHMKM", "PROT_DISJ_POLOS",
        "PROT_DISJ_TENSAO_V", "PROT_DISJ_CORRENTE_A", "PROT_DISJ_FREQ_HZ",
        "PROT_DISJ_CAP_INT_KA", "PROT_DISJ_CURVA", "DSV_POLOS", "DSV_TENSAO_V",
        "DSV_CORRENTE_A", "DSV_FECHAMENTO_CC_A", "DSV_CAP_INT_KA", "DSV_CURVA",
        "ATERRAMENTO_ESQUEMA", "ATERRAMENTO_RESISTENCIA_OHM", "ATERRAMENTO_HASTES_QTD",
        "ATERRAMENTO_HASTE_M", "ATERRAMENTO_HASTE_SECAO", "ATERRAMENTO_BITOLA_MM2",
        "CABO_ISOLACAO", "CABO_ISOLAMENTO_KV", "CABO_BITOLA_MM2", "CABO_CAPACIDADE_A",
        "CABO_METODO_INSTALACAO", "FATOR_TEMPERATURA", "FATOR_AGRUPAMENTO",
        "TEMPERATURA_AMBIENTE_C", "DIAG_DISJ_INV_A",
    ]),
]


class MemorialApp(tk.Tk):
    def __init__(self) -> None:
        super().__init__()
        self.title("Gerador de Memorial Solar")
        self.geometry("1180x780")
        self.minsize(980, 650)

        user_dir = Path.home() / "Documents" / "Gerador Memorial Solar"
        user_dir.mkdir(parents=True, exist_ok=True)
        self.user_dir = user_dir
        self.project_db = user_dir / "historico_projetos.sqlite3"
        user_excel = user_dir / "Mapa_Dados_Memorial_LIGHT_100_PARAM.xlsx"
        template_excel = resource_path("Mapa_Dados_Memorial_LIGHT_100_PARAM.xlsx")
        if not user_excel.exists():
            shutil.copy2(template_excel, user_excel)
        else:
            ensure_workbook_schema(user_excel, template_excel)

        self.excel = tk.StringVar(value=str(user_excel))
        self.template = tk.StringVar(value=str(resource_path("TEMPLATE_MEMORIAL_LIGHT_100_PARAM_TAGUEADO.docx")))
        self.output = tk.StringVar(value=str(user_dir / "Memorial_Gerado.docx"))
        self.documents: list[str] = []
        self.consumption_suggestions: list[int] = []
        self.module_variants: list[dict[str, str]] = []
        self.inverter_variants: list[dict[str, str]] = []
        self.fields: dict[str, tk.StringVar] = {}
        self.consumption_fields: dict[int, tk.StringVar] = {}
        self.load_fields: dict[tuple[int, str], tk.StringVar] = {}
        self.concessionaire = tk.StringVar(value="LIGHT")
        self.compensation_mode = tk.StringVar(value="Compensação local")
        self.remainder_installation = tk.StringVar()
        self.form_options: dict[str, tk.BooleanVar] = {}
        self.allocation_rows: list[dict[str, tk.StringVar]] = []
        self.diagram_enabled = tk.BooleanVar(value=True)
        self._build()
        self.load_excel()

    def _path_row(self, parent, row: int, label: str, variable: tk.StringVar, command) -> None:
        ttk.Label(parent, text=label).grid(row=row, column=0, sticky="w", padx=6, pady=4)
        ttk.Entry(parent, textvariable=variable).grid(row=row, column=1, sticky="ew", padx=6, pady=4)
        ttk.Button(parent, text="Selecionar", command=command).grid(row=row, column=2, padx=6, pady=4)

    def _scroll_tab(self, title: str) -> ttk.Frame:
        outer = ttk.Frame(self.notebook)
        self.notebook.add(outer, text=title)
        canvas = tk.Canvas(outer, highlightthickness=0)
        scrollbar = ttk.Scrollbar(outer, orient="vertical", command=canvas.yview)
        frame = ttk.Frame(canvas)
        frame.bind("<Configure>", lambda event: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.create_window((0, 0), window=frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        return frame

    def _build(self) -> None:
        top = ttk.LabelFrame(self, text="Arquivos do projeto")
        top.pack(fill="x", padx=12, pady=10)
        top.columnconfigure(1, weight=1)
        self._path_row(top, 0, "Planilha", self.excel, self.choose_excel)
        self._path_row(top, 1, "Template Word", self.template, self.choose_template)
        self._path_row(top, 2, "Memorial de saida", self.output, self.choose_output)

        actions = ttk.Frame(self)
        actions.pack(fill="x", padx=12)
        ttk.Button(actions, text="Adicionar documentos", command=self.choose_documents).pack(side="left", padx=4)
        ttk.Button(actions, text="Analisar documentos", command=self.analyze_documents).pack(side="left", padx=4)
        ttk.Button(actions, text="Novo projeto", command=self.new_project).pack(side="left", padx=4)
        ttk.Button(actions, text="Recarregar planilha", command=self.load_excel).pack(side="left", padx=4)
        ttk.Button(actions, text="Salvar dados", command=self.save_excel).pack(side="left", padx=4)
        ttk.Checkbutton(
            actions,
            text="Gerar diagrama unifilar (PDF + DXF)",
            variable=self.diagram_enabled,
        ).pack(side="left", padx=12)
        ttk.Button(actions, text="Gerar memorial", command=self.generate).pack(side="right", padx=4)

        self.status = tk.StringVar(value="O programa aceita PDF, DOCX e imagens. Revise os dados detectados antes de gerar.")
        ttk.Label(self, textvariable=self.status, foreground="#204a87").pack(fill="x", padx=16, pady=8)

        detected = ttk.LabelFrame(self, text="Equipamentos encontrados nos datasheets")
        detected.pack(fill="x", padx=12, pady=(0, 8))
        detected.columnconfigure(1, weight=1)
        ttk.Label(detected, text="Modulo").grid(row=0, column=0, sticky="w", padx=6, pady=4)
        self.module_choice = tk.StringVar()
        self.module_combo = ttk.Combobox(
            detected, textvariable=self.module_choice, state="disabled", width=68
        )
        self.module_combo.grid(row=0, column=1, sticky="ew", padx=6, pady=4)
        self.module_combo.bind("<<ComboboxSelected>>", self.apply_module_variant)
        ttk.Label(detected, text="Inversor").grid(row=1, column=0, sticky="w", padx=6, pady=4)
        self.inverter_choice = tk.StringVar()
        self.inverter_combo = ttk.Combobox(
            detected, textvariable=self.inverter_choice, state="disabled", width=68
        )
        self.inverter_combo.grid(row=1, column=1, sticky="ew", padx=6, pady=4)
        self.inverter_combo.bind("<<ComboboxSelected>>", self.apply_inverter_variant)

        self.notebook = ttk.Notebook(self)
        self.notebook.pack(fill="both", expand=True, padx=12, pady=8)
        self.inputs_frame = self._scroll_tab("Formulario")
        self.consumption_frame = self._scroll_tab("Consumo 12 meses")
        self.loads_frame = self._scroll_tab("Cargas da residencia")
        self.forms_frame = self._scroll_tab("Concessionária e rateio")
        self._build_forms_tab()

    def _build_forms_tab(self) -> None:
        frame = self.forms_frame
        ttk.Label(frame, text="Concessionária", font=("Segoe UI", 10, "bold")).grid(
            row=0, column=0, sticky="w", padx=6, pady=(10, 4)
        )
        combo = ttk.Combobox(
            frame,
            textvariable=self.concessionaire,
            values=["LIGHT", "Enel-RJ"],
            state="readonly",
            width=32,
        )
        combo.grid(row=0, column=1, sticky="w", padx=6, pady=(10, 4))
        combo.bind("<<ComboboxSelected>>", self._update_concessionaire_options)

        ttk.Label(frame, text="Modalidade de compensação").grid(
            row=1, column=0, sticky="w", padx=6, pady=4
        )
        ttk.Combobox(
            frame,
            textvariable=self.compensation_mode,
            values=[
                "Compensação local",
                "Autoconsumo remoto",
                "Múltiplas unidades consumidoras",
                "Geração compartilhada",
            ],
            state="readonly",
            width=38,
        ).grid(row=1, column=1, sticky="w", padx=6, pady=4)

        forms_box = ttk.LabelFrame(frame, text="Formulários a gerar")
        forms_box.grid(row=2, column=0, columnspan=5, sticky="ew", padx=6, pady=8)
        options = [
            ("connection", "Solicitação para orçamento de conexão", True),
            ("generator", "Registro da central geradora", True),
            ("gd_data", "Dados GD - UFV", True),
            ("priority", "Rateio por ordem de prioridade", False),
            ("percentage", "Rateio por percentuais", False),
        ]
        for row, (key, label, default) in enumerate(options):
            variable = tk.BooleanVar(value=default)
            self.form_options[key] = variable
            ttk.Checkbutton(forms_box, text=label, variable=variable).grid(
                row=row, column=0, sticky="w", padx=8, pady=3
            )

        ttk.Label(
            frame,
            text="Contas participantes do rateio (a unidade geradora será prioridade 0)",
            font=("Segoe UI", 10, "bold"),
        ).grid(row=3, column=0, columnspan=5, sticky="w", padx=6, pady=(12, 4))
        self.allocation_container = ttk.Frame(frame)
        self.allocation_container.grid(row=4, column=0, columnspan=5, sticky="ew", padx=6)
        headers = ["Prioridade", "Instalação/UC", "Código cliente", "Titular", "Endereço", "Conta"]
        for column, label in enumerate(headers):
            ttk.Label(
                self.allocation_container, text=label, font=("Segoe UI", 9, "bold")
            ).grid(row=0, column=column, sticky="w", padx=3, pady=3)
        self.add_allocation_row()

        controls = ttk.Frame(frame)
        controls.grid(row=5, column=0, columnspan=5, sticky="w", padx=6, pady=8)
        ttk.Button(controls, text="Adicionar outra conta", command=self.add_allocation_row).pack(
            side="left", padx=3
        )
        ttk.Button(controls, text="Remover última conta", command=self.remove_allocation_row).pack(
            side="left", padx=3
        )
        ttk.Label(controls, text="Destino de eventual saldo:").pack(side="left", padx=(18, 4))
        ttk.Entry(controls, textvariable=self.remainder_installation, width=24).pack(side="left")

        ttk.Label(
            frame,
            text=(
                "Na LIGHT, o programa preenche os modelos fornecidos. "
                "Na Enel-RJ, gera um pacote de dados para apoiar o protocolo no formulário/portal vigente."
            ),
            foreground="#7a4b00",
            wraplength=900,
        ).grid(row=6, column=0, columnspan=5, sticky="w", padx=6, pady=8)

    def _update_concessionaire_options(self, _event=None) -> None:
        is_light = self.concessionaire.get() == "LIGHT"
        template_name = MEMORIAL_TEMPLATES.get(
            self.concessionaire.get(), MEMORIAL_TEMPLATES["LIGHT"]
        )
        self.template.set(str(resource_path(template_name)))
        for key in ("connection", "generator", "gd_data", "percentage"):
            if not is_light:
                self.form_options[key].set(False)
        if not is_light:
            self.form_options["priority"].set(bool(self.allocation_rows))
        self.status.set(
            "LIGHT: modelos oficiais disponíveis."
            if is_light
            else "Enel-RJ: será gerado um pacote de dados para conferência e protocolo."
        )

    def add_allocation_row(self) -> None:
        index = len(self.allocation_rows) + 1
        row = {
            "installation": tk.StringVar(),
            "customer_code": tk.StringVar(),
            "holder": tk.StringVar(),
            "address": tk.StringVar(),
            "percentage": tk.StringVar(),
            "file": tk.StringVar(),
        }
        self.allocation_rows.append(row)
        grid_row = len(self.allocation_rows)
        ttk.Label(self.allocation_container, text=str(index)).grid(
            row=grid_row, column=0, padx=3, pady=3
        )
        widths = [18, 16, 24, 42]
        for column, (key, width) in enumerate(
            zip(("installation", "customer_code", "holder", "address"), widths), start=1
        ):
            ttk.Entry(
                self.allocation_container, textvariable=row[key], width=width
            ).grid(row=grid_row, column=column, sticky="ew", padx=3, pady=3)
        ttk.Button(
            self.allocation_container,
            text="Ler conta",
            command=lambda current=row: self.choose_allocation_bill(current),
        ).grid(row=grid_row, column=5, padx=3, pady=3)

    def remove_allocation_row(self) -> None:
        if not self.allocation_rows:
            return
        self.allocation_rows.pop()
        grid_row = len(self.allocation_rows) + 1
        for widget in self.allocation_container.grid_slaves(row=grid_row):
            widget.destroy()

    def choose_allocation_bill(self, row: dict[str, tk.StringVar]) -> None:
        selected = filedialog.askopenfilename(
            filetypes=[("Conta de energia", "*.pdf *.png *.jpg *.jpeg"), ("Todos", "*.*")]
        )
        if not selected:
            return
        row["file"].set(selected)
        try:
            suggestions, unsupported, _consumption, _choices = suggest_fields([selected])
        except Exception as exc:
            messagebox.showerror("Conta para rateio", str(exc))
            return
        row["installation"].set(suggestions.get("UC_CONTA_CONTRATO", ""))
        row["customer_code"].set(suggestions.get("UC_CODIGO_CLIENTE", ""))
        row["holder"].set(suggestions.get("CLIENTE_NOME", ""))
        address = " ".join(
            value
            for value in (
                suggestions.get("CLIENTE_ENDERECO_LOGRADOURO", ""),
                suggestions.get("CLIENTE_BAIRRO_MUN_UF_CEP", ""),
            )
            if value
        )
        row["address"].set(address)
        if unsupported:
            self.status.set("A conta foi anexada, mas a leitura automática não encontrou dados. Preencha a linha.")
        else:
            self.status.set("Conta de rateio analisada. Confira instalação, titular e endereço.")

    def new_project(self) -> None:
        if not messagebox.askyesno(
            "Novo projeto",
            "Criar um novo projeto? O projeto atual sera salvo no historico e a planilha sera copiada para a pasta Backups antes de limpar os campos.",
        ):
            return
        self.save_excel()
        project_name = _project_label({key: var.get().strip() for key, var in self.fields.items()})
        backup = _backup_workbook(Path(self.excel.get()), project_name)
        clear_prefixes = ("CLIENTE_", "UC_", "COORD_", "FV_", "MOD_", "INV_", "INMETRO_", "COTACAO_", "STRING_", "RT_ART_TRT")
        keep_keys = {"CONCESSIONARIA", "ESTADO", "RT_NOME", "RT_CPF", "RT_REGISTRO", "RT_EMAIL", "RT_CELULAR", "TIPO_RAMAL", "TIPO_SOLICITACAO"}
        for key, variable in self.fields.items():
            if key in keep_keys:
                continue
            if key.startswith(clear_prefixes) or key in {"PROJETO_NOME", "PROJETO_STATUS", "DATA_PROJETO", "LOCAL_CIDADE_UF"}:
                variable.set("")
        self.fields.get("PROJETO_STATUS", tk.StringVar()).set("Em preenchimento")
        for variable in self.consumption_fields.values():
            variable.set("")
        self.module_variants = []
        self.inverter_variants = []
        self._configure_variant_combo(self.module_combo, self.module_choice, [])
        self._configure_variant_combo(self.inverter_combo, self.inverter_choice, [])
        self.documents = []
        self.save_excel(make_backup=False, update_history=False)
        msg = "Novo projeto criado. Campos antigos foram arquivados no historico."
        if backup:
            msg += f" Backup: {backup}"
        self.status.set(msg)


    def choose_excel(self) -> None:
        path = filedialog.askopenfilename(filetypes=[("Excel", "*.xlsx")])
        if path:
            self.excel.set(path)
            self.load_excel()

    def choose_template(self) -> None:
        path = filedialog.askopenfilename(filetypes=[("Word", "*.docx")])
        if path:
            self.template.set(path)

    def choose_output(self) -> None:
        path = filedialog.asksaveasfilename(defaultextension=".docx", filetypes=[("Word", "*.docx")])
        if path:
            self.output.set(path)

    def choose_documents(self) -> None:
        paths = filedialog.askopenfilenames(
            filetypes=[("Documentos", "*.pdf *.docx *.txt *.csv *.png *.jpg *.jpeg"), ("Todos", "*.*")]
        )
        self.documents = list(paths)
        self.status.set(f"{len(self.documents)} documento(s) selecionado(s).")

    def _clear_frames(self) -> None:
        for frame in (self.inputs_frame, self.consumption_frame, self.loads_frame):
            for widget in frame.winfo_children():
                widget.destroy()

    def load_excel(self) -> None:
        try:
            workbook = load_workbook(self.excel.get(), data_only=False)
        except Exception as exc:
            messagebox.showerror("Planilha", str(exc))
            return

        self._clear_frames()
        self.fields.clear()
        self.consumption_fields.clear()
        self.load_fields.clear()
        inputs = workbook["Inputs"]
        input_values = {
            str(inputs[f"A{row}"].value or ""): inputs[f"B{row}"].value
            for row in range(2, inputs.max_row + 1)
        }
        for key, value in EXTRA_INPUT_DEFAULTS.items():
            input_values.setdefault(key, value)

        display_row = 0
        used: set[str] = set()
        for group, keys in FIELD_GROUPS:
            ttk.Label(self.inputs_frame, text=group, font=("Segoe UI", 10, "bold")).grid(
                row=display_row, column=0, columnspan=2, sticky="w", padx=6, pady=(10, 4)
            )
            display_row += 1
            for key in keys:
                value = input_values.get(key)
                variable = tk.StringVar(value="" if value is None else str(value))
                self.fields[key] = variable
                used.add(key)
                ttk.Label(self.inputs_frame, text=key, width=36).grid(row=display_row, column=0, sticky="w", padx=6, pady=3)
                ttk.Entry(self.inputs_frame, textvariable=variable, width=72).grid(
                    row=display_row, column=1, sticky="ew", padx=6, pady=3
                )
                display_row += 1
        for key, value in input_values.items():
            if key and key not in used:
                variable = tk.StringVar(value="" if value is None else str(value))
                self.fields[key] = variable
                ttk.Label(self.inputs_frame, text=key, width=36).grid(row=display_row, column=0, sticky="w", padx=6, pady=3)
                ttk.Entry(self.inputs_frame, textvariable=variable, width=72).grid(
                    row=display_row, column=1, sticky="ew", padx=6, pady=3
                )
                display_row += 1
        self.inputs_frame.columnconfigure(1, weight=1)

        consumo = workbook["Consumo_24m"]
        ttk.Label(self.consumption_frame, text="Preencha os 12 meses da conta de energia", font=("Segoe UI", 10, "bold")).grid(
            row=0, column=0, columnspan=2, sticky="w", padx=6, pady=(10, 4)
        )
        for idx in range(1, 13):
            value = consumo[f"B{idx + 1}"].value
            variable = tk.StringVar(value="" if value is None else str(value))
            self.consumption_fields[idx] = variable
            ttk.Label(self.consumption_frame, text=f"Consumo mes {idx} (kWh)", width=28).grid(
                row=idx, column=0, sticky="w", padx=6, pady=3
            )
            ttk.Entry(self.consumption_frame, textvariable=variable, width=24).grid(
                row=idx, column=1, sticky="w", padx=6, pady=3
            )

        cargas = workbook["Cargas_24"]
        headers = [("DESC", "Descricao"), ("P_W", "Potencia W"), ("QTD", "Quantidade"), ("FP", "FP")]
        for col, (_, label) in enumerate(headers, start=1):
            ttk.Label(self.loads_frame, text=label, font=("Segoe UI", 9, "bold")).grid(row=0, column=col, padx=6, pady=4)
        for idx in range(1, 13):
            ttk.Label(self.loads_frame, text=f"Carga {idx}").grid(row=idx, column=0, sticky="w", padx=6, pady=3)
            for col_offset, (suffix, _) in enumerate(headers, start=2):
                value = cargas.cell(idx + 1, col_offset).value
                variable = tk.StringVar(value="" if value is None else str(value))
                self.load_fields[(idx, suffix)] = variable
                ttk.Entry(self.loads_frame, textvariable=variable, width=24).grid(
                    row=idx, column=col_offset - 1, sticky="ew", padx=4, pady=3
                )

    def analyze_documents(self) -> None:
        if not self.documents:
            messagebox.showwarning("Documentos", "Adicione os PDFs ou documentos antes de analisar.")
            return
        try:
            suggestions, unsupported, consumption, choices = suggest_fields(self.documents)
        except Exception as exc:
            messagebox.showerror("Analise", str(exc))
            return
        suggestions = expand_shared_values(suggestions)
        for key, value in suggestions.items():
            if key in self.fields and not self.fields[key].get().strip():
                self.fields[key].set(value)
        if self.fields.get("CONCESSIONARIA") and self.fields["CONCESSIONARIA"].get().strip():
            concessionaire = self.fields["CONCESSIONARIA"].get().strip().upper()
            self.concessionaire.set("Enel-RJ" if "ENEL" in concessionaire else "LIGHT")
            self._update_concessionaire_options()
        self.consumption_suggestions = consumption
        for idx, value in enumerate(consumption[:12], start=1):
            if idx in self.consumption_fields and not self.consumption_fields[idx].get().strip():
                self.consumption_fields[idx].set(str(value))
        self.module_variants = choices["modules"]
        self.inverter_variants = choices["inverters"]
        self._configure_variant_combo(self.module_combo, self.module_choice, self.module_variants)
        self._configure_variant_combo(self.inverter_combo, self.inverter_choice, self.inverter_variants)
        note = f"{len(suggestions)} campo(s) sugerido(s)."
        if consumption:
            note += f" {len(consumption)} consumo(s) mensais detectado(s)."
        if self.module_variants:
            note += f" {len(self.module_variants)} modulo(s) para escolher."
        if self.inverter_variants:
            note += f" {len(self.inverter_variants)} inversor(es) para escolher."
        if unsupported:
            note += " Nao foi possivel ler: " + ", ".join(unsupported)
        self.status.set(note + " Revise tudo antes de gerar.")

    @staticmethod
    def _configure_variant_combo(
        combo: ttk.Combobox,
        variable: tk.StringVar,
        variants: list[dict[str, str]],
    ) -> None:
        labels = [variant["label"] for variant in variants]
        combo.configure(values=labels, state="readonly" if labels else "disabled")
        if labels:
            variable.set("Selecione um modelo...")
        else:
            variable.set("")

    def _apply_variant(self, variants: list[dict[str, str]], selected: str) -> None:
        variant = next((item for item in variants if item["label"] == selected), None)
        if not variant:
            return
        for key, value in variant.items():
            if key != "label" and key in self.fields:
                self.fields[key].set(value)
        self.status.set(f"Dados de {selected} aplicados ao formulario. Revise antes de gerar.")

    def apply_module_variant(self, _event=None) -> None:
        self._apply_variant(self.module_variants, self.module_choice.get())

    def apply_inverter_variant(self, _event=None) -> None:
        self._apply_variant(self.inverter_variants, self.inverter_choice.get())

    def save_excel(self, make_backup: bool = True, update_history: bool = True) -> None:
        try:
            current_values = {
                key: variable.get().strip() for key, variable in self.fields.items()
            }
            current_values["CONCESSIONARIA"] = self.concessionaire.get()
            if not current_values.get("DATA_PROJETO"):
                current_values["DATA_PROJETO"] = date.today().strftime("%d/%m/%Y")
            shared_values = expand_shared_values(current_values)
            for key, value in shared_values.items():
                if key in self.fields and not self.fields[key].get().strip() and value not in (None, ""):
                    self.fields[key].set(str(value))
                    current_values[key] = str(value)
            has_project_data = any(
                current_values.get(key)
                for key in ("PROJETO_NOME", "CLIENTE_NOME", "UC_CONTA_CONTRATO", "FV_POT_KWP", "COTACAO_NUMERO")
            )
            if make_backup and has_project_data:
                _backup_workbook(Path(self.excel.get()), _project_label(current_values))
            workbook = load_workbook(self.excel.get(), data_only=False)
            inputs = workbook["Inputs"]
            existing = {
                str(inputs[f"A{row}"].value or "") for row in range(2, inputs.max_row + 1)
            }
            for key in self.fields:
                if key not in existing:
                    dest = inputs.max_row + 1
                    inputs[f"A{dest}"] = key
                    existing.add(key)
            for row in range(2, inputs.max_row + 1):
                key = str(inputs[f"A{row}"].value or "")
                if key in self.fields:
                    inputs[f"B{row}"] = self.fields[key].get().strip() or None

            consumo = workbook["Consumo_24m"]
            for idx, variable in self.consumption_fields.items():
                consumo[f"B{idx + 1}"] = variable.get().strip() or None

            cargas = workbook["Cargas_24"]
            column_by_suffix = {"DESC": 2, "P_W": 3, "QTD": 4, "FP": 5}
            load_values = {}
            for (idx, suffix), variable in self.load_fields.items():
                value = variable.get().strip()
                load_values[(idx, suffix)] = value
                cargas.cell(idx + 1, column_by_suffix[suffix]).value = value or None

            consumption_values = {idx: variable.get().strip() for idx, variable in self.consumption_fields.items()}
            if update_history and has_project_data:
                _append_history_row(workbook, current_values)
            workbook.save(self.excel.get())
            if update_history and has_project_data:
                _save_project_database(self.project_db, current_values, consumption_values, load_values, self.documents)
            self.status.set("Dados salvos na planilha e no historico do projeto." if has_project_data else "Dados salvos na planilha.")
        except Exception as exc:
            messagebox.showerror("Salvar", str(exc))

    def generate(self) -> None:
        self.save_excel()
        try:
            result = generate_memorial(self.excel.get(), self.template.get(), self.output.get())
        except Exception as exc:
            messagebox.showerror("Geracao", str(exc))
            return
        selected_forms = {key for key, variable in self.form_options.items() if variable.get()}
        accounts = [
            AllocationAccount(
                installation=row["installation"].get().strip(),
                customer_code=row["customer_code"].get().strip(),
                holder=row["holder"].get().strip(),
                address=row["address"].get().strip(),
                percentage=row["percentage"].get().strip(),
            )
            for row in self.allocation_rows
            if any(
                row[key].get().strip()
                for key in ("installation", "customer_code", "holder", "address")
            )
        ]
        if accounts:
            selected_forms.add("priority")
        package_dir = Path(self.output.get()).with_suffix("").parent / (
            Path(self.output.get()).stem + "_Formularios"
        )
        try:
            generated_forms = generate_concessionaire_package(
                self.excel.get(),
                resource_path("forms") / "light",
                package_dir,
                self.concessionaire.get(),
                selected_forms,
                self.compensation_mode.get(),
                accounts,
                self.remainder_installation.get().strip(),
            )
        except Exception as exc:
            messagebox.showerror(
                "Formulários",
                f"O memorial foi gerado, mas ocorreu um erro nos formulários:\n{exc}",
            )
            generated_forms = []
        details = [f"Memorial gerado em:\n{result.output}"]
        if generated_forms:
            details.append(f"\n{len(generated_forms)} formulário(s) gerado(s) em:\n{package_dir}")
        if self.diagram_enabled.get():
            diagram_base = Path(self.output.get()).with_suffix("").parent / (
                Path(self.output.get()).stem + "_Diagrama_Unifilar"
            )
            try:
                diagram_template = DIAGRAM_TEMPLATES.get(
                    self.concessionaire.get(), DIAGRAM_TEMPLATES["LIGHT"]
                )
                diagram = generate_single_line_diagram(
                    self.excel.get(),
                    resource_path("cad") / diagram_template,
                    diagram_base,
                )
                details.append(
                    f"\nDiagrama unifilar gerado com {len(diagram.strings)} string(s) "
                    f"({', '.join(str(value) for value in diagram.strings)} módulos):"
                    f"\n{diagram.pdf}\n{diagram.dxf}"
                )
                if diagram.warnings:
                    details.append("\nAlertas do diagrama:\n" + "\n".join(diagram.warnings))
            except Exception as exc:
                details.append(f"\nO diagrama não foi gerado:\n{exc}")
        if result.missing_values:
            details.append(f"\nAtencao: {len(result.missing_values)} campos estavam vazios.")
        if result.unresolved_tags:
            details.append("\nTags nao resolvidas:\n" + ", ".join(result.unresolved_tags))
        messagebox.showinfo("Concluido", "\n".join(details))
        self.status.set("Geracao concluida. Revise tecnicamente o documento antes de protocolar.")


if __name__ == "__main__":
    MemorialApp().mainloop()
