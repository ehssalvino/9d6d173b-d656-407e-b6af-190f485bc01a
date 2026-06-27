from __future__ import annotations

import math
import re
from dataclasses import dataclass
from pathlib import Path
from typing import Any

import ezdxf
import fitz
import matplotlib

matplotlib.use("Agg")
import matplotlib.backends.backend_pdf  # noqa: F401 - required by PyInstaller PDF export
import matplotlib.pyplot as plt
from ezdxf import bbox
from ezdxf.addons.drawing import Frontend, RenderContext
from ezdxf.addons.drawing.config import BackgroundPolicy, ColorPolicy, Configuration
from ezdxf.addons.drawing.matplotlib import MatplotlibBackend
from openpyxl import load_workbook
from PIL import Image

from .core import _make_location_figure
from .shared_data import expand_shared_values
from .string_design import StringConfig, design_string_arrangement


class DiagramError(RuntimeError):
    pass


@dataclass
class DiagramResult:
    dxf: Path
    pdf: Path
    strings: list[int]
    warnings: list[str]


STANDARD_BREAKERS = [10, 16, 20, 25, 32, 40, 50, 63, 70, 80, 100, 125]
DRAWING_LIMITS = (100.0, 950.0, -8070.0, -7440.0)


def _text(value: Any) -> str:
    return "" if value is None else str(value).strip()


def _number(value: Any) -> float | None:
    if value in (None, ""):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace(" ", "")
    match = re.search(r"-?\d+(?:[.,]\d+)?", text)
    if not match:
        return None
    try:
        return float(match.group(0).replace(",", "."))
    except ValueError:
        return None


def _fmt(value: float | int | None, decimals: int = 2) -> str:
    if value is None:
        return "—"
    if isinstance(value, int) or float(value).is_integer():
        return str(int(value))
    return f"{float(value):.{decimals}f}".replace(".", ",")


def _inputs(excel_path: str | Path) -> dict[str, Any]:
    workbook = load_workbook(excel_path, data_only=False)
    sheet = workbook["Inputs"]
    values = {
        _text(sheet[f"A{row}"].value): sheet[f"B{row}"].value
        for row in range(2, sheet.max_row + 1)
        if sheet[f"A{row}"].value
    }
    return expand_shared_values(values)


def _breaker(value: Any, nominal_current: float | None) -> int:
    specified = _number(value)
    target = specified or ((nominal_current or 0) * 1.25)
    for rating in STANDARD_BREAKERS:
        if rating >= target:
            return rating
    return math.ceil(target / 25) * 25


def _phase_conductors(connection: str) -> int:
    normalized = connection.upper()
    if "TRI" in normalized:
        return 4
    if "BI" in normalized:
        return 3
    return 2


def _connection_poles(connection: str) -> int:
    normalized = connection.upper()
    if "TRI" in normalized:
        return 3
    if "BI" in normalized:
        return 2
    return 1


def _conductor_notation(connection: str, section_mm2: float | None) -> str:
    normalized = connection.upper()
    section = _fmt(section_mm2, 0)
    if "TRI" in normalized:
        return f"3F+N+T #{section} mm²"
    if "BI" in normalized:
        return f"2F+N+T #{section} mm²"
    return f"F+N+T #{section} mm²"


def _breaker_kind(connection: str) -> str:
    normalized = connection.upper()
    if "TRI" in normalized:
        return "DISJUNTOR GERAL TRIPOLAR"
    if "BI" in normalized:
        return "DISJUNTOR GERAL BIPOLAR"
    return "DISJUNTOR GERAL MONOPOLAR"


def _design_strings(data: dict[str, Any]) -> tuple[list[StringConfig], list[str]]:
    try:
        design = design_string_arrangement(data, strict=True)
    except ValueError as exc:
        raise DiagramError(str(exc)) from exc
    return design.strings, design.warnings


def _string_summary(configs: list[StringConfig]) -> tuple[str, str, str]:
    modules = [item.modules for item in configs]
    if len(configs) == 1:
        return "ARRANJO SOLAR - 1 STRING", f"STRING {configs[0].label} DE {configs[0].modules} PAINÉIS", f"CC {configs[0].label}"
    labels = "; ".join(f"CC {item.label}" for item in configs)
    if len(set(modules)) == 1:
        return f"ARRANJO SOLAR - {len(configs)} STRINGS", f"{len(configs)} STRINGS DE {modules[0]} PAINÉIS", labels
    return (
        f"ARRANJO SOLAR - {len(configs)} STRINGS",
        "STRINGS DE " + " + ".join(f"{item.modules} ({item.label})" for item in configs) + " PAINÉIS",
        labels,
    )


def _draw_string_list(modelspace, configs: list[StringConfig], x: float = 604.0, y: float = -7725.0) -> None:
    for index, item in enumerate(configs):
        _put(modelspace, x, y - index * 7.0, f"CC {item.label} - {item.modules} módulos", radius=4, height=2.0)


def _entity_text(entity) -> str:
    if entity.dxftype() == "TEXT":
        return entity.dxf.text
    if entity.dxftype() == "MTEXT":
        return entity.text
    return ""


def _set_entity_text(entity, value: str) -> None:
    if entity.dxftype() == "TEXT":
        entity.dxf.text = value
        entity.dxf.style = "Style-Arial"
    elif entity.dxftype() == "MTEXT":
        entity.text = value
        entity.dxf.style = "Style-Arial"


def _insert(entity) -> tuple[float, float]:
    point = entity.dxf.insert
    return float(point.x), float(point.y)


def _nearest_text(modelspace, x: float, y: float, radius: float = 8.0):
    candidates = []
    for entity in modelspace.query("TEXT MTEXT"):
        px, py = _insert(entity)
        distance = math.hypot(px - x, py - y)
        if distance <= radius:
            candidates.append((distance, entity))
    if not candidates:
        return None
    return min(candidates, key=lambda item: item[0])[1]


def _add_text(modelspace, x: float, y: float, value: str, height: float = 2.8) -> None:
    if "\n" in value or len(value) > 80:
        mtext = modelspace.add_mtext(
            value,
            dxfattribs={
                "insert": (x, y),
                "char_height": height,
                "style": "Style-Arial",
                "color": 7,
            },
        )
        mtext.dxf.width = 150
        return
    modelspace.add_text(
        value,
        dxfattribs={
            "insert": (x, y),
            "height": height,
            "style": "Style-Arial",
            "color": 7,
        },
    )


def _put(
    modelspace,
    x: float,
    y: float,
    value: str,
    radius: float = 8.0,
    create: bool = True,
    height: float = 2.8,
) -> None:
    entity = _nearest_text(modelspace, x, y, radius)
    if entity is not None:
        _set_entity_text(entity, value)
    elif create:
        _add_text(modelspace, x, y, value, height)


def _delete_text_if(modelspace, predicate) -> None:
    for entity in list(modelspace.query("TEXT MTEXT")):
        value = _entity_text(entity)
        if predicate(value, *_insert(entity)):
            modelspace.delete_entity(entity)


def _clear_dynamic_diagram_text(modelspace) -> None:
    dynamic_terms = (
        "arranjo solar",
        "string",
        "painéis",
        "paineis",
        "max potência",
        "max potencia",
        "voc ",
        "isc ",
        "fabricante:",
        "modelo:",
        "rede ",
        "disjuntor geral",
        "potência nominal ca",
        "potencia nominal ca",
        "dados técnicos",
        "dados tecnicos",
        "registro do inmetro",
        "máx.",
        "max.",
        "min.",
        "tensão",
        "tensao",
        "potência",
        "potencia",
        "corrente",
        "frequ",
        "distor",
        "fator de pot",
        "dimens",
        "peso:",
        "grau de prote",
        "classe de prote",
        "consumo noturno",
        "topologia",
        "controle de temperatura",
        "temperatura ambiente",
        "umidade",
        "humidade",
        "conexões",
        "conexoes",
        "bitola",
        "módulo",
        "modulo",
        " wp",
        "placa solar",
        "emissão automática",
        "emissao automatica",
        "diagrama_",
    )
    left_cable_terms = ("rede", "pvc", "xlpe", "classe 2", "mm²", "mm2")

    def should_delete(value: str, x: float, y: float) -> bool:
        normalized = value.lower()
        in_dynamic_area = 130 <= x <= 910 and -8040 <= y <= -7480
        if not in_dynamic_area:
            return False
        if re.fullmatch(r"\s*\d+\s*n\s*\d+(?:[,.]\d+)?\s*mm2\s*", normalized):
            return True
        if re.fullmatch(r"\s*\d+\s*a\s*", normalized):
            return True
        if x < 540 and any(term in normalized for term in left_cable_terms):
            return True
        return any(term in normalized for term in dynamic_terms)

    _delete_text_if(modelspace, should_delete)


def _clean_template(modelspace) -> None:
    xmin, xmax, ymin, ymax = DRAWING_LIMITS
    for entity in list(modelspace):
        if entity.dxftype() == "OLE2FRAME":
            modelspace.delete_entity(entity)
            continue
        try:
            extent = bbox.extents([entity], fast=True)
        except Exception:
            continue
        if not extent.has_data:
            continue
        center = extent.center
        if center.x < xmin - 100 or center.x > xmax + 100 or center.y < ymin - 100 or center.y > ymax + 100:
            modelspace.delete_entity(entity)


def _remove_second_string(modelspace) -> None:
    for entity in list(modelspace):
        try:
            extent = bbox.extents([entity], fast=True)
        except Exception:
            continue
        if not extent.has_data:
            continue
        center = extent.center
        if 545 <= center.x <= 710 and -7720 <= center.y <= -7637:
            modelspace.delete_entity(entity)


def _replace_utility_notes(modelspace, concessionaire: str) -> None:
    is_light = "LIGHT" in concessionaire.upper()
    for entity in modelspace.query("TEXT MTEXT"):
        value = _entity_text(entity)
        if not value:
            continue
        if is_light:
            continue
        value = re.sub(r"\bLIGHT\b", concessionaire, value, flags=re.IGNORECASE)
        value = value.replace(
            "RECON-BT - " + concessionaire,
            "normas técnicas vigentes da concessionária",
        )
        _set_entity_text(entity, value)


def _fill_drawing(modelspace, data: dict[str, Any], configs: list[StringConfig]) -> None:
    _clear_dynamic_diagram_text(modelspace)
    strings = [item.modules for item in configs]
    qty = sum(strings)
    module_power = _number(data.get("MOD_WP")) or 0
    voc = _number(data.get("MOD_VOC")) or 0
    isc = _number(data.get("MOD_ISC")) or 0
    max_string = max(strings)
    total_power_w = qty * module_power
    inverter_power_w = (_number(data.get("INV_PN_KW")) or 0) * 1000
    voltage_ac = _number(data.get("INV_VAC_NOM")) or 220
    connection = _text(data.get("TIPO_LIGACAO")) or "BIFÁSICO"
    phases = _phase_conductors(connection)
    poles = int(_number(data.get("DISJ_POLOS")) or _connection_poles(connection))
    nominal_current = _number(data.get("INV_IMAX_CA"))
    if nominal_current is None and inverter_power_w:
        nominal_current = (
            inverter_power_w / (math.sqrt(3) * voltage_ac)
            if "TRI" in connection.upper()
            else inverter_power_w / voltage_ac
        )
    inverter_breaker = _breaker(data.get("DIAG_DISJ_INV_A") or data.get("PROT_CA"), nominal_current)
    main_breaker = int(_number(data.get("DISJ_CORRENTE_A")) or 0)
    entry_cable = _number(data.get("CONDUTOR_FASE_MM2"))
    ac_cable = _number(data.get("BITOLA_CA_MM2") or data.get("CABO_BITOLA_MM2"))
    cc_cable = _number(data.get("BITOLA_CC_MM2") or data.get("CABO_BITOLA_MM2"))
    cable_class = _text(data.get("CABO_ISOLACAO")) or "XLPE"
    cable_voltage = _text(data.get("CABO_ISOLAMENTO_KV")) or "0,6/1"
    concessionaire = _text(data.get("CONCESSIONARIA")) or "LIGHT"

    arrangement_title, arrangement_modules, arrangement_labels = _string_summary(configs)
    current_arrangement_labels = (
        arrangement_labels
        if len(configs) <= 4
        else f"{len(configs)} strings ({configs[0].label} a {configs[-1].label})"
    )
    module_type = _text(data.get("MOD_TIPO_CELULA")) or "FOTOVOLTAICOS"
    _put(modelspace, 410.4, -7491.3, arrangement_title)
    _put(modelspace, 401.8, -7498.6, f"{arrangement_modules} {module_type}")
    _put(modelspace, 388.3, -7507.0, f"MAX POTÊNCIA {_fmt(total_power_w, 0)} W,")
    _put(modelspace, 456.7, -7506.6, f"Voc {_fmt(max_string * voc)} V,")
    _put(modelspace, 496.2, -7506.4, f"Isc {_fmt(isc)} A")
    _put(
        modelspace,
        372.7,
        -7516.0,
        f"FABRICANTE: {_text(data.get('MOD_MARCA'))} MODELO: {_text(data.get('MOD_MODELO'))} "
        f"{_fmt(module_power, 0)} W",
        radius=12,
    )
    _put(modelspace, 148.8, -7606.2, f"REDE {concessionaire}", height=2.4)
    _put(modelspace, 254.0, -7585.2, _breaker_kind(connection), radius=12, height=2.4)
    _put(modelspace, 263.6, -7599.1, f"{poles}P - {main_breaker} A")
    _put(modelspace, 276.3, -7607.8, _conductor_notation(connection, entry_cable))
    _put(
        modelspace,
        168.9,
        -7623.1,
        f"{_conductor_notation(connection, entry_cable)} - {cable_class} - {cable_voltage} kV",
        radius=14,
    )
    _put(modelspace, 483.8, -7586.9, f"{_connection_poles(connection)}P - {inverter_breaker} A")
    _put(modelspace, 501.4, -7588.1, _conductor_notation(connection, ac_cable))
    _put(modelspace, 402.3, -7588.1, _conductor_notation(connection, ac_cable))
    _put(
        modelspace,
        548.0,
        -7558.0,
        f"INVERSOR 01\nPOTÊNCIA NOMINAL CA: {_fmt(inverter_power_w, 0)} W\n"
        f"FABRICANTE: {_text(data.get('INV_MARCA'))}\nMODELO: {_text(data.get('INV_MODELO'))}",
        radius=15,
        height=2.0,
    )
    _put(modelspace, 640.5, -7598.0, f"{_fmt(module_power, 0)} Wp")
    _put(modelspace, 641.7, -7627.3, f"{_fmt(module_power, 0)} Wp")
    _put(modelspace, 647.6, -7671.8, f"{_fmt(module_power, 0)} Wp")
    _put(modelspace, 648.8, -7701.0, f"{_fmt(module_power, 0)} Wp")
    _put(modelspace, 637.1, -7575.9, f"CC {configs[0].label} módulo 01")
    _put(modelspace, 639.6, -7607.5, f"módulo {strings[0]:02d}")
    if len(strings) >= 2:
        _put(modelspace, 644.1, -7649.6, f"CC {configs[1].label} módulo 01")
        _put(modelspace, 646.7, -7681.3, f"módulo {strings[1]:02d}")
    else:
        _remove_second_string(modelspace)
    if len(strings) > 2:
        _draw_string_list(modelspace, configs)


    technical = {
        (154.7, -7654.9): f"DADOS TÉCNICOS DO INVERSOR {_text(data.get('INV_MARCA'))} {_text(data.get('INV_MODELO'))}",
        (154.7, -7661.5): f"REGISTRO DO INMETRO {_text(data.get('INMETRO_REGISTRO'))} {_text(data.get('INMETRO_DATA'))}",
        (154.7, -7680.5): f"Máx. corrente de alimentação CC: {_fmt(_number(data.get('INV_ICC_MAX')))} A",
        (154.7, -7687.1): f"Isc por string: {_fmt(isc)} A; arranjo: {current_arrangement_labels}",
        (154.7, -7693.6): f"Mín. tensão de entrada: {_fmt(_number(data.get('INV_VMPPT_MIN')))} V",
        (155.2, -7700.2): f"Tensão de partida: {_fmt(_number(data.get('INV_VSTART')))} V",
        (155.2, -7706.7): f"Tensão CC máxima: {_fmt(_number(data.get('INV_VCC_MAX')))} V",
        (154.7, -7713.3): f"Faixa MPPT: {_fmt(_number(data.get('INV_VMPPT_MIN')))} a {_fmt(_number(data.get('INV_VMPPT_MAX')))} V",
        (154.8, -7719.2): "",
        (154.7, -7725.1): f"Número de MPPTs: {_fmt(_number(data.get('INV_MPPTS')), 0)}",
        (154.7, -7730.9): f"Número de conexões CC: {_fmt(_number(data.get('INV_STRINGS')), 0)}",
        (154.7, -7736.8): f"Máx. potência do gerador FV: {_fmt(_number(data.get('INV_PMAXCC_KW')))} kWp",
        (154.7, -7755.2): f"Potência nominal CA: {_fmt(inverter_power_w, 0)} W",
        (154.7, -7761.0): f"Potência máxima CA: {_fmt((_number(data.get('INV_PCA_MAX_KW')) or 0) * 1000, 0)} VA",
        (154.8, -7766.9): f"Corrente de saída nominal: {_fmt(nominal_current)} A",
        (155.2, -7772.8): f"Tensão de conexão: {_fmt(voltage_ac, 0)} V {connection.title()}",
        (154.7, -7778.7): f"Frequência: {_fmt(_number(data.get('INV_FN')), 0)} Hz",
        (154.7, -7784.6): f"Distorção harmônica total: {_text(data.get('INV_THD'))} %",
        (154.7, -7791.1): f"Fator de potência: {_text(data.get('INV_FP'))}",
        (154.7, -7809.5): f"Dimensões: {_text(data.get('INV_DIMENSOES')) or '330 x 310 x 172 mm'}",
        (154.7, -7815.4): f"Peso: {_text(data.get('INV_PESO_KG')) or '11'} kg",
        (154.8, -7821.2): f"Grau de proteção: {_text(data.get('INV_IP')) or 'IP65'}",
        (154.8, -7827.1): f"Classe de proteção: {_text(data.get('INV_CLASSE_PROTECAO')) or 'Classe I'}",
        (154.8, -7833.0): f"Consumo noturno: {_text(data.get('INV_CONSUMO_NOTURNO_W')) or '< 1'} W",
        (155.2, -7838.9): f"Topologia: {_text(data.get('INV_TOPOLOGIA')) or 'Sem transformador'}",
        (154.8, -7844.8): f"Controle de temperatura: {_text(data.get('INV_REFRIGERACAO')) or 'Refrigeração natural'}",
        (154.8, -7850.7): f"Temperatura ambiente: {_text(data.get('INV_TEMPERATURA')) or '-25 a +60 °C'}",
        (154.7, -7856.5): f"Umidade relativa do ar: {_text(data.get('INV_UMIDADE')) or '0 a 100%'}",
        (154.8, -7862.4): f"Conexões com o arranjo solar: {_fmt(_number(data.get('INV_STRINGS')), 0)} CC+ e CC-",
        (154.7, -7869.2): f"Bitola dos terminais CC: {_fmt(cc_cable, 0)} mm²",
        (154.8, -7875.1): f"Conexão com a rede: {_connection_poles(connection)} polos CA + N + T",
        (154.7, -7881.9): f"Bitola dos terminais CA: {_fmt(ac_cable, 0)} mm²",
    }
    for (x, y), value in technical.items():
        _put(modelspace, x, y, value, radius=9)

    address = " ".join(
        item
        for item in (
            _text(data.get("CLIENTE_ENDERECO_LOGRADOURO")),
            _text(data.get("CLIENTE_BAIRRO_MUN_UF_CEP")),
        )
        if item
    )
    _put(modelspace, 779.2, -7914.0, _text(data.get("LOCAL_CIDADE_UF")))
    _put(modelspace, 775.5, -7920.3, address, radius=10)
    _put(modelspace, 783.7, -7927.2, f"{_fmt(total_power_w / 1000)} kWp")
    _put(modelspace, 884.2, -7927.2, _text(data.get("UC_CONTA_CONTRATO")))
    _put(modelspace, 800.9, -7958.8, _text(data.get("CLIENTE_NOME")))
    _put(modelspace, 870.3, -7958.8, _text(data.get("CLIENTE_CPF")))
    _put(modelspace, 804.9, -7984.9, _text(data.get("RT_NOME")))
    _put(modelspace, 870.3, -7984.9, _text(data.get("RT_CPF")))
    _put(modelspace, 833.8, -8027.0, f"DIAGRAMA_{_text(data.get('CLIENTE_NOME')).replace(' ', '_')}")
    _put(modelspace, 765.0, -7858.5, "EMISSÃO AUTOMÁTICA")
    _put(modelspace, 909.8, -7858.6, _text(data.get("DATA_PROJETO")))
    _put(modelspace, 770.1, -7658.4, f"PLACA SOLAR {_text(data.get('MOD_MARCA'))}")
    _put(modelspace, 770.8, -7661.7, f"{_text(data.get('MOD_MODELO'))} {_fmt(module_power, 0)} W")
    _put(modelspace, 770.0, -7706.0, f"{_text(data.get('INV_MARCA'))} - {_text(data.get('INV_MODELO'))}")
    _replace_utility_notes(modelspace, concessionaire)
    for entity in modelspace.query("TEXT MTEXT"):
        entity.dxf.color = 7
        if entity.dxf.hasattr("transparency"):
            entity.dxf.discard("transparency")


def _render_pdf(
    document,
    output: Path,
    data: dict[str, Any],
    map_path: Path | None,
) -> None:
    xmin, xmax, ymin, ymax = DRAWING_LIMITS
    figure = plt.figure(figsize=(16.54, 11.69), dpi=150)
    axes = figure.add_axes([0.015, 0.015, 0.97, 0.97])
    context = RenderContext(document)
    backend = MatplotlibBackend(axes)
    configuration = Configuration(
        color_policy=ColorPolicy.BLACK,
        background_policy=BackgroundPolicy.WHITE,
    )
    Frontend(context, backend, config=configuration).draw_layout(
        document.modelspace(), finalize=True
    )
    figure.set_size_inches(16.54, 11.69, forward=True)
    if map_path and map_path.exists():
        with Image.open(map_path) as image:
            axes.imshow(
                image.convert("RGB"),
                extent=(545, 744, -8035, -7890),
                aspect="auto",
                zorder=0,
            )
    axes.set_xlim(xmin, xmax)
    axes.set_ylim(ymin, ymax)
    axes.set_aspect("equal", adjustable="box")
    axes.axis("off")
    output.parent.mkdir(parents=True, exist_ok=True)
    figure.savefig(output, format="pdf", facecolor="white")
    plt.close(figure)


def generate_single_line_diagram(
    excel_path: str | Path,
    template_path: str | Path,
    output_base: str | Path,
) -> DiagramResult:
    data = _inputs(excel_path)
    configs, warnings = _design_strings(data)
    strings = [item.modules for item in configs]
    template = Path(template_path)
    if not template.exists():
        raise DiagramError(f"Template do diagrama não encontrado: {template}")

    base = Path(output_base)
    if base.suffix:
        base = base.with_suffix("")
    dxf_output = base.with_suffix(".dxf")
    pdf_output = base.with_suffix(".pdf")
    dxf_output.parent.mkdir(parents=True, exist_ok=True)

    document = ezdxf.readfile(template)
    modelspace = document.modelspace()
    _clean_template(modelspace)
    _fill_drawing(modelspace, data, configs)
    for layer in document.layers:
        layer.dxf.color = 7
        if layer.dxf.hasattr("transparency"):
            layer.dxf.discard("transparency")
    document.saveas(dxf_output)

    map_path = base.parent / f"{base.name}_mapa.png"
    address = " ".join(
        item
        for item in (
            _text(data.get("CLIENTE_ENDERECO_LOGRADOURO")),
            _text(data.get("CLIENTE_BAIRRO_MUN_UF_CEP")),
        )
        if item
    )
    try:
        _make_location_figure(
            data.get("COORD_SUL"),
            data.get("COORD_OESTE"),
            address,
            map_path,
        )
    except Exception:
        map_path = None

    _render_pdf(document, pdf_output, data, map_path)
    if map_path and map_path.exists():
        map_path.unlink(missing_ok=True)
    return DiagramResult(dxf_output, pdf_output, strings, warnings)
